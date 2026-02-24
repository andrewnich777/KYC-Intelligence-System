"""
Excel Export Generator.

Generates a multi-sheet Excel workbook (.xlsx) with:
  1. Executive Summary
  2. Screening Results
  3. Risk Factors
  4. UBO/Ownership (business clients)
  5. Decision Points
  6. Regulatory Actions
  7. Evidence Detail

Uses openpyxl for workbook generation with conditional formatting,
auto-filters, frozen headers, and clickable hyperlinks.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from constants import FAILED_SENTINEL_KEY
from logger import get_logger

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
RISK_FILLS = {
    "LOW": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),       # Green
    "MEDIUM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),     # Yellow
    "HIGH": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),       # Orange-red
    "CRITICAL": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),   # Red
}

RISK_FONTS = {
    "LOW": Font(color="006100", bold=True),
    "MEDIUM": Font(color="9C5700", bold=True),
    "HIGH": Font(color="9C0006", bold=True),
    "CRITICAL": Font(color="FFFFFF", bold=True),
}

HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
HEADER_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

LINK_FONT = Font(color="0563C1", underline="single")

DISP_CLEAR_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
DISP_MATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

EVIDENCE_LEVEL_FILLS = {
    "V": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "S": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "I": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "U": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}


def _style_header_row(ws: Any, col_count: int) -> None:
    """Apply header styling + freeze + auto-filter to row 1."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _auto_fit_columns(ws: Any, min_width: int = 10, max_width: int = 50) -> None:
    """Set column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value or "")
            # Use first line only for multi-line cells
            first_line = val.split("\n")[0] if "\n" in val else val
            max_len = max(max_len, len(first_line))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _write_row(ws: Any, row_idx: int, values: list, styles: dict | None = None) -> None:
    """Write a row of values with optional per-cell styles."""
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = CELL_ALIGNMENT
        cell.border = THIN_BORDER
        if styles:
            for attr, style_val in styles.items():
                setattr(cell, attr, style_val)


def _safe_str(value: Any, max_len: int = 0) -> str:
    """Safely convert a value to string, optionally truncating."""
    s = str(value) if value is not None else ""
    if max_len and len(s) > max_len:
        return s[:max_len] + "..."
    return s


def _apply_risk_style(cell: Any, risk_level: str) -> None:
    """Apply conditional formatting based on risk level."""
    level = risk_level.upper() if isinstance(risk_level, str) else ""
    if level in RISK_FILLS:
        cell.fill = RISK_FILLS[level]
        cell.font = RISK_FONTS[level]


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_executive_summary(wb: Workbook, output: Any) -> None:
    """Sheet 1: Executive Summary."""
    ws = wb.active
    ws.title = "Executive Summary"

    # Key-value layout
    fields: list[tuple[str, str]] = []

    client_data = output.client_data or {}
    client_name = client_data.get("full_name") or client_data.get("legal_name", "Unknown")
    fields.append(("Client Name", client_name))
    fields.append(("Client Type", output.client_type.value.upper()))
    fields.append(("Client ID", output.client_id))

    # Risk
    risk_level = "N/A"
    risk_score = "N/A"
    if output.synthesis and output.synthesis.revised_risk_assessment:
        ra = output.synthesis.revised_risk_assessment
        risk_level = ra.risk_level.value
        risk_score = str(ra.total_score)
    elif output.intake_classification and output.intake_classification.preliminary_risk:
        ra = output.intake_classification.preliminary_risk
        risk_level = ra.risk_level.value
        risk_score = str(ra.total_score)
    fields.append(("Risk Level", risk_level))
    fields.append(("Risk Score", risk_score))

    # Decision
    decision = output.final_decision.value if output.final_decision else "PENDING"
    fields.append(("Decision", decision))

    # Key findings
    if output.synthesis and output.synthesis.key_findings:
        fields.append(("Key Findings Count", str(len(output.synthesis.key_findings))))
        for i, finding in enumerate(output.synthesis.key_findings[:5], 1):
            fields.append((f"  Finding {i}", _safe_str(finding, 120)))

    # Evidence quality
    if output.review_intelligence and output.review_intelligence.confidence:
        conf = output.review_intelligence.confidence
        grade = conf.overall_confidence_grade
        pcts = f"V:{conf.verified_pct:.0f}% S:{conf.sourced_pct:.0f}% I:{conf.inferred_pct:.0f}% U:{conf.unknown_pct:.0f}%"
        fields.append(("Evidence Quality Grade", grade))
        fields.append(("Evidence Breakdown", pcts))
        if conf.degraded:
            fields.append(("Evidence Status", "DEGRADED"))

    # Review
    if output.review_session:
        fields.append(("Analyst", output.review_session.officer_name or "Not specified"))
        fields.append(("Review Finalized", str(output.review_session.finalized)))
    fields.append(("Generated At", output.generated_at.strftime("%Y-%m-%d %H:%M:%S")))
    fields.append(("Duration (seconds)", f"{output.duration_seconds:.1f}"))

    if output.is_degraded:
        fields.append(("Pipeline Status", "DEGRADED (some agents failed)"))

    # Write
    ws.append(["Field", "Value"])
    _style_header_row(ws, 2)
    for row_idx, (field_name, field_val) in enumerate(fields, start=2):
        ws.cell(row=row_idx, column=1, value=field_name).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).alignment = CELL_ALIGNMENT
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        val_cell = ws.cell(row=row_idx, column=2, value=field_val)
        val_cell.alignment = CELL_ALIGNMENT
        val_cell.border = THIN_BORDER

        # Apply risk colour to risk level cell
        if field_name == "Risk Level":
            _apply_risk_style(val_cell, field_val)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80


def _build_screening_results(wb: Workbook, output: Any) -> None:
    """Sheet 2: Screening Results — one row per evidence record with screening data."""
    ws = wb.create_sheet("Screening Results")

    headers = [
        "Entity Screened", "Check Type", "Disposition", "Confidence",
        "Source", "Source Tier", "Key Finding", "Evidence ID", "Source URLs",
    ]
    ws.append(headers)

    investigation = output.investigation_results
    if not investigation:
        _style_header_row(ws, len(headers))
        _auto_fit_columns(ws)
        return

    row_idx = 2

    def _add_evidence_rows(check_type: str, result: Any) -> int:
        nonlocal row_idx
        if not result or not hasattr(result, "evidence_records"):
            return row_idx
        for er in result.evidence_records:
            urls = ", ".join(er.source_urls) if er.source_urls else ""
            supporting = er.supporting_data or []
            tier = ""
            for sd in supporting:
                if isinstance(sd, dict) and sd.get("source_tier"):
                    tier = sd["source_tier"]
                    break
            values = [
                er.entity_screened,
                check_type,
                er.disposition.value if hasattr(er.disposition, "value") else str(er.disposition),
                er.confidence.value if hasattr(er.confidence, "value") else str(er.confidence),
                er.source_name,
                tier,
                _safe_str(er.claim, 100),
                er.evidence_id,
                urls,
            ]
            _write_row(ws, row_idx, values)

            # Style disposition cell
            disp_str = values[2]
            disp_cell = ws.cell(row=row_idx, column=3)
            if disp_str == "CLEAR":
                disp_cell.fill = DISP_CLEAR_FILL
            elif disp_str in ("POTENTIAL_MATCH", "CONFIRMED_MATCH"):
                disp_cell.fill = DISP_MATCH_FILL

            # Make URLs clickable
            if er.source_urls:
                url_cell = ws.cell(row=row_idx, column=9)
                first_url = er.source_urls[0]
                if first_url.startswith("https://") or first_url.startswith("http://"):
                    url_cell.hyperlink = first_url
                    url_cell.font = LINK_FONT

            row_idx += 1
        return row_idx

    _add_evidence_rows("Sanctions (Individual)", investigation.individual_sanctions)
    _add_evidence_rows("Sanctions (Entity)", investigation.entity_sanctions)
    _add_evidence_rows("PEP Detection", investigation.pep_classification)
    _add_evidence_rows("Adverse Media (Individual)", investigation.individual_adverse_media)
    _add_evidence_rows("Adverse Media (Business)", investigation.business_adverse_media)
    _add_evidence_rows("Entity Verification", investigation.entity_verification)
    _add_evidence_rows("Jurisdiction Risk", investigation.jurisdiction_risk)
    _add_evidence_rows("Transaction Monitoring", investigation.transaction_monitoring)

    # UBO screening
    for ubo_name, ubo_data in (investigation.ubo_screening or {}).items():
        if not isinstance(ubo_data, dict):
            continue
        for check_key in ("sanctions", "pep", "adverse_media"):
            check_result = ubo_data.get(check_key)
            if check_result and isinstance(check_result, dict):
                for er_dict in check_result.get("evidence_records", []):
                    entity = er_dict.get("entity_screened", ubo_name)
                    urls = ", ".join(er_dict.get("source_urls", []))
                    values = [
                        entity,
                        f"UBO {check_key.replace('_', ' ').title()}",
                        er_dict.get("disposition", "PENDING_REVIEW"),
                        er_dict.get("confidence", "MEDIUM"),
                        er_dict.get("source_name", ""),
                        "",
                        _safe_str(er_dict.get("claim", ""), 100),
                        er_dict.get("evidence_id", ""),
                        urls,
                    ]
                    _write_row(ws, row_idx, values)
                    row_idx += 1

    _style_header_row(ws, len(headers))
    _auto_fit_columns(ws)


def _build_risk_factors(wb: Workbook, output: Any) -> None:
    """Sheet 3: Risk Factors breakdown."""
    ws = wb.create_sheet("Risk Factors")

    headers = ["Factor", "Points", "Category", "Source", "Description"]
    ws.append(headers)

    risk_assessment = None
    if output.synthesis and output.synthesis.revised_risk_assessment:
        risk_assessment = output.synthesis.revised_risk_assessment
    elif output.intake_classification and output.intake_classification.preliminary_risk:
        risk_assessment = output.intake_classification.preliminary_risk

    row_idx = 2
    if risk_assessment and risk_assessment.risk_factors:
        # Sort by points descending
        sorted_factors = sorted(risk_assessment.risk_factors, key=lambda f: f.points, reverse=True)
        for rf in sorted_factors:
            values = [rf.factor, rf.points, rf.category, rf.source, ""]
            _write_row(ws, row_idx, values)
            # Colour-code high-point factors
            pts_cell = ws.cell(row=row_idx, column=2)
            if rf.points >= 15:
                pts_cell.fill = RISK_FILLS["HIGH"]
                pts_cell.font = RISK_FONTS["HIGH"]
            elif rf.points >= 10:
                pts_cell.fill = RISK_FILLS["MEDIUM"]
                pts_cell.font = RISK_FONTS["MEDIUM"]
            row_idx += 1

        # Total row
        total_score = risk_assessment.total_score
        ws.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=total_score).font = Font(bold=True)
        _apply_risk_style(ws.cell(row=row_idx, column=2), risk_assessment.risk_level.value)

    _style_header_row(ws, len(headers))
    _auto_fit_columns(ws)


def _build_ubo_ownership(wb: Workbook, output: Any) -> None:
    """Sheet 4: UBO/Ownership (business clients only)."""
    investigation = output.investigation_results
    client_data = output.client_data or {}

    # Only for business clients
    if output.client_type.value != "business":
        return
    if not client_data.get("beneficial_owners") and not (investigation and investigation.ubo_screening):
        return

    ws = wb.create_sheet("UBO Ownership")

    headers = [
        "Owner Name", "Ownership %", "Nationality", "PEP Status",
        "Sanctions Status", "Adverse Media", "Risk Notes",
    ]
    ws.append(headers)

    row_idx = 2

    # From client data
    for bo in client_data.get("beneficial_owners", []):
        name = bo.get("full_name", "Unknown")
        pct = bo.get("ownership_percentage", "?")
        nationality = bo.get("citizenship", "N/A")

        # Enrich from UBO screening
        pep_status = "Not screened"
        sanctions_status = "Not screened"
        media_status = "Not screened"
        notes = ""

        if investigation and investigation.ubo_screening:
            ubo_data = investigation.ubo_screening.get(name, {})
            if isinstance(ubo_data, dict):
                pep_result = ubo_data.get("pep", {})
                if isinstance(pep_result, dict):
                    if pep_result.get(FAILED_SENTINEL_KEY):
                        pep_status = "ERROR"
                    else:
                        pep_status = pep_result.get("detected_level", "NOT_PEP")
                sanctions_result = ubo_data.get("sanctions", {})
                if isinstance(sanctions_result, dict):
                    if sanctions_result.get(FAILED_SENTINEL_KEY):
                        sanctions_status = "ERROR"
                    else:
                        sanctions_status = sanctions_result.get("disposition", "CLEAR")
                media_result = ubo_data.get("adverse_media", {})
                if isinstance(media_result, dict):
                    if media_result.get(FAILED_SENTINEL_KEY):
                        media_status = "ERROR"
                    else:
                        media_status = media_result.get("overall_level", "CLEAR")

        if bo.get("pep_self_declaration"):
            notes = "Self-declared PEP"

        values = [name, pct, nationality, pep_status, sanctions_status, media_status, notes]
        _write_row(ws, row_idx, values)

        # Colour PEP cell
        pep_cell = ws.cell(row=row_idx, column=4)
        if pep_status not in ("NOT_PEP", "Not screened"):
            pep_cell.fill = RISK_FILLS["HIGH"]
            pep_cell.font = RISK_FONTS["HIGH"]

        row_idx += 1

    _style_header_row(ws, len(headers))
    _auto_fit_columns(ws)


def _build_decision_points(wb: Workbook, output: Any) -> None:
    """Sheet 5: Decision Points from synthesis + review session."""
    ws = wb.create_sheet("Decision Points")

    headers = [
        "Decision ID", "Topic", "Recommendation", "Confidence",
        "Officer Decision", "Officer Notes", "Counter-Argument",
    ]
    ws.append(headers)

    row_idx = 2
    if output.synthesis and output.synthesis.decision_points:
        for dp in output.synthesis.decision_points:
            officer_decision = ""
            if dp.officer_selection:
                for opt in dp.options:
                    if opt.option_id == dp.officer_selection:
                        officer_decision = opt.label
                        break
                if not officer_decision:
                    officer_decision = dp.officer_selection

            values = [
                dp.decision_id,
                dp.title,
                dp.disposition,
                f"{dp.confidence:.0%}",
                officer_decision,
                dp.officer_notes or "",
                dp.counter_argument.argument if dp.counter_argument else "",
            ]
            _write_row(ws, row_idx, values)
            row_idx += 1

    _style_header_row(ws, len(headers))
    _auto_fit_columns(ws)


def _build_regulatory_actions(wb: Workbook, output: Any) -> None:
    """Sheet 6: Regulatory Actions from compliance_actions + review_intelligence."""
    ws = wb.create_sheet("Regulatory Actions")

    headers = [
        "Action Type", "Obligation", "Deadline", "Regulatory Basis",
        "Status", "Filing Details",
    ]
    ws.append(headers)

    row_idx = 2

    # From investigation compliance_actions
    investigation = output.investigation_results
    if investigation and investigation.compliance_actions:
        ca = investigation.compliance_actions
        for action in ca.get("actions", []):
            if isinstance(action, str):
                values = ["Action", action, "", "", "PENDING", ""]
            elif isinstance(action, dict):
                values = [
                    action.get("action_type", ""),
                    action.get("description", ""),
                    action.get("deadline", ""),
                    action.get("regulatory_basis", ""),
                    action.get("status", "PENDING"),
                    action.get("details", ""),
                ]
            else:
                continue
            _write_row(ws, row_idx, values)
            row_idx += 1

        for filing in ca.get("reports", []):
            values = [
                "Filing",
                filing.get("type", ""),
                filing.get("timeline", ""),
                filing.get("regulatory_basis", ""),
                filing.get("filing_decision", "PENDING"),
                filing.get("details", ""),
            ]
            _write_row(ws, row_idx, values)
            row_idx += 1

    # From review intelligence regulatory mappings
    if output.review_intelligence and output.review_intelligence.regulatory_mappings:
        for fm in output.review_intelligence.regulatory_mappings:
            for tag in fm.regulatory_tags:
                values = [
                    "Regulatory Tag",
                    tag.obligation,
                    tag.timeline,
                    tag.regulation,
                    "Filing Required" if tag.filing_required else "Advisory",
                    tag.trigger_description,
                ]
                _write_row(ws, row_idx, values)
                row_idx += 1

    _style_header_row(ws, len(headers))
    _auto_fit_columns(ws)


def _build_evidence_detail(wb: Workbook, output: Any, evidence_store: list | None = None) -> None:
    """Sheet 7: Full evidence detail — the drill-down sheet."""
    ws = wb.create_sheet("Evidence Detail")

    headers = [
        "Evidence ID", "Source Type", "Source Name", "Agent Name",
        "Entity Screened", "Entity Context", "Claim",
        "Evidence Level", "Disposition", "Disposition Reasoning",
        "Confidence", "Timestamp", "Source URLs",
        "Search Queries",
    ]
    ws.append(headers)

    records = evidence_store or []
    row_idx = 2

    for er in records:
        if isinstance(er, dict):
            urls = er.get("source_urls", [])
            urls_str = "\n".join(urls) if urls else ""

            # Extract search queries from supporting_data
            queries = []
            for sd in (er.get("supporting_data") or []):
                if isinstance(sd, dict) and sd.get("search_query"):
                    queries.append(sd["search_query"])

            values = [
                er.get("evidence_id", ""),
                er.get("source_type", ""),
                er.get("source_name", ""),
                er.get("agent_name", "") or er.get("source_type", ""),
                er.get("entity_screened", ""),
                er.get("entity_context", ""),
                _safe_str(er.get("claim", ""), 200),
                er.get("evidence_level", "U"),
                er.get("disposition", "PENDING_REVIEW"),
                _safe_str(er.get("disposition_reasoning", ""), 150),
                er.get("confidence", "MEDIUM"),
                er.get("timestamp", ""),
                urls_str,
                "; ".join(queries),
            ]
            _write_row(ws, row_idx, values)

            # Evidence level colour coding
            level = er.get("evidence_level", "U")
            level_cell = ws.cell(row=row_idx, column=8)
            if level in EVIDENCE_LEVEL_FILLS:
                level_cell.fill = EVIDENCE_LEVEL_FILLS[level]

            # Make first URL clickable
            if urls:
                url_cell = ws.cell(row=row_idx, column=13)
                if urls[0].startswith("https://") or urls[0].startswith("http://"):
                    url_cell.hyperlink = urls[0]
                    url_cell.font = LINK_FONT

            row_idx += 1

    _style_header_row(ws, len(headers))
    _auto_fit_columns(ws)


def _build_filing_worksheet(
    wb: Workbook,
    title: str,
    filing: dict | None,
) -> None:
    """Build a human-readable filing worksheet from a nested filing dict.

    Renders a flat Field / Value / Section layout so analysts can review
    and edit pre-filled SAR/STR fields directly in Excel.
    """
    if not filing:
        return

    ws = wb.create_sheet(title)
    ws.append(["Section", "Field", "Value"])

    row_idx = 2

    def _flatten(obj: Any, section: str, depth: int = 0) -> None:
        nonlocal row_idx
        if isinstance(obj, dict):
            for key, val in obj.items():
                if key in ("filing_notes",):
                    # Render notes as separate rows
                    for i, note in enumerate(val if isinstance(val, list) else [val], 1):
                        _write_row(ws, row_idx, [section, f"Note {i}", str(note)])
                        row_idx += 1
                elif isinstance(val, dict):
                    _flatten(val, section=key.replace("_", " ").title(), depth=depth + 1)
                elif isinstance(val, list):
                    if val and isinstance(val[0], dict):
                        for i, item in enumerate(val, 1):
                            _flatten(item, section=f"{key.replace('_', ' ').title()} #{i}", depth=depth + 1)
                    else:
                        _write_row(ws, row_idx, [section, key.replace("_", " ").title(), ", ".join(str(v) for v in val)])
                        row_idx += 1
                else:
                    _write_row(ws, row_idx, [section, key.replace("_", " ").title(), str(val) if val is not None else ""])
                    row_idx += 1

    # Top-level metadata
    for meta_key in ("form", "schema_version", "generated_at", "client_id"):
        if meta_key in filing:
            _write_row(ws, row_idx, ["Metadata", meta_key.replace("_", " ").title(), str(filing[meta_key])])
            row_idx += 1

    # Parts
    for key, val in filing.items():
        if key.startswith("part_") and isinstance(val, dict):
            section_name = key.replace("_", " ").title()
            _flatten(val, section=section_name)

    # Filing notes
    for note in filing.get("filing_notes", []):
        _write_row(ws, row_idx, ["Filing Notes", "", str(note)])
        ws.cell(row=row_idx, column=3).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        row_idx += 1

    _style_header_row(ws, 3)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 80


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_excel(
    output: Any,
    output_path: Path | None = None,
    evidence_store: list | None = None,
    fincen_filing: dict | None = None,
    fintrac_filing: dict | None = None,
) -> Path:
    """Generate a multi-sheet Excel workbook from KYCOutput.

    Args:
        output: KYCOutput instance.
        output_path: Where to save the .xlsx file. If None, derives from output_dir.
        evidence_store: Flat list of evidence record dicts.
        fincen_filing: Pre-generated FinCEN SAR filing dict for worksheet sheet.
        fintrac_filing: Pre-generated FINTRAC STR filing dict for worksheet sheet.

    Returns:
        Path to the generated .xlsx file.
    """
    wb = Workbook()

    _build_executive_summary(wb, output)
    _build_screening_results(wb, output)
    _build_risk_factors(wb, output)
    _build_ubo_ownership(wb, output)
    _build_decision_points(wb, output)
    _build_regulatory_actions(wb, output)
    _build_evidence_detail(wb, output, evidence_store)
    _build_filing_worksheet(wb, "SAR Filing Worksheet", fincen_filing)
    _build_filing_worksheet(wb, "STR Filing Worksheet", fintrac_filing)

    # Determine output path
    if output_path is None:
        output_path = Path("results") / output.client_id / "05_output" / "screening_results.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Excel workbook saved: %s", output_path)
    return output_path
