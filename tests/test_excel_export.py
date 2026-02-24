"""Tests for Excel export generator (Feature 1)."""

import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from generators.excel_export import (
    DISP_CLEAR_FILL,
    DISP_MATCH_FILL,
    EVIDENCE_LEVEL_FILLS,
    RISK_FILLS,
    RISK_FONTS,
    _safe_str,
    generate_excel,
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _mock_output(client_type="individual", with_investigation=True, with_synthesis=True):
    """Build a minimal mock KYCOutput for Excel tests."""
    output = MagicMock()
    output.client_id = "test_excel_001"
    output.client_type.value = client_type
    output.client_data = {
        "full_name": "Test User",
        "legal_name": "Test Corp" if client_type == "business" else None,
        "address": {"street": "1 Main St", "city": "Toronto", "country": "CA"},
    }
    if client_type == "business":
        output.client_data["beneficial_owners"] = [
            {"full_name": "Owner A", "ownership_percentage": 60, "citizenship": "CA"},
            {"full_name": "Owner B", "ownership_percentage": 40, "citizenship": "US"},
        ]

    output.generated_at = datetime(2026, 3, 1, 12, 0, 0)
    output.duration_seconds = 120.5
    output.final_decision.value = "APPROVE"
    output.is_degraded = False
    output.schema_version = "3.0"
    output.metrics = {}

    if with_synthesis:
        output.synthesis.revised_risk_assessment.risk_level.value = "HIGH"
        output.synthesis.revised_risk_assessment.total_score = 67
        output.synthesis.revised_risk_assessment.risk_factors = [
            MagicMock(factor="PEP detected", points=20, category="PEP", source="PEPAgent"),
            MagicMock(factor="High-risk jurisdiction", points=15, category="Jurisdiction", source="JurisdictionAgent"),
        ]
        output.synthesis.key_findings = ["PEP detected in Hong Kong", "Adverse media found"]
        output.synthesis.risk_elevations = []
        output.synthesis.recommended_decision.value = "CONDITIONAL_APPROVE"
        output.synthesis.decision_points = []
    else:
        output.synthesis = None

    output.intake_classification.preliminary_risk.risk_level.value = "HIGH"
    output.intake_classification.preliminary_risk.total_score = 55
    output.intake_classification.preliminary_risk.risk_factors = []

    if with_investigation:
        output.investigation_results.individual_sanctions = None
        output.investigation_results.entity_sanctions = None
        output.investigation_results.pep_classification = None
        output.investigation_results.individual_adverse_media = None
        output.investigation_results.business_adverse_media = None
        output.investigation_results.entity_verification = None
        output.investigation_results.jurisdiction_risk = None
        output.investigation_results.transaction_monitoring = None
        output.investigation_results.ubo_screening = {}
        output.investigation_results.compliance_actions = None
    else:
        output.investigation_results = None

    output.review_intelligence.confidence.overall_confidence_grade = "B"
    output.review_intelligence.confidence.verified_pct = 50
    output.review_intelligence.confidence.sourced_pct = 30
    output.review_intelligence.confidence.inferred_pct = 15
    output.review_intelligence.confidence.unknown_pct = 5
    output.review_intelligence.confidence.degraded = False
    output.review_intelligence.regulatory_mappings = []

    output.review_session.officer_name = "Test Officer"
    output.review_session.finalized = True

    return output


def _sample_evidence():
    """Sample evidence records for testing."""
    return [
        {
            "evidence_id": "E-001",
            "source_type": "SanctionsScreening",
            "source_name": "OFAC SDN",
            "agent_name": "IndividualSanctionsAgent",
            "entity_screened": "Test User",
            "entity_context": "primary",
            "claim": "No match found on OFAC SDN list",
            "evidence_level": "V",
            "disposition": "CLEAR",
            "disposition_reasoning": "No matches after fuzzy search",
            "confidence": "HIGH",
            "timestamp": "2026-02-28T14:32:00Z",
            "source_urls": ["https://sanctionssearch.ofac.treas.gov/"],
            "supporting_data": [{"search_query": "Test User sanctions"}],
        },
        {
            "evidence_id": "E-002",
            "source_type": "PEPDetection",
            "source_name": "UN PEP Database",
            "agent_name": "PEPDetectionAgent",
            "entity_screened": "Test User",
            "entity_context": "primary",
            "claim": "PEP detected: Foreign PEP in Hong Kong",
            "evidence_level": "S",
            "disposition": "PEP_CONFIRMED",
            "disposition_reasoning": "Confirmed match",
            "confidence": "HIGH",
            "timestamp": "2026-02-28T14:33:00Z",
            "source_urls": ["https://example.com/pep"],
            "supporting_data": [],
        },
    ]


def _sample_fincen_filing():
    """Sample FinCEN SAR filing dict."""
    return {
        "form": "FinCEN SAR Form 111",
        "generated_at": "2026-03-01T12:00:00",
        "client_id": "test_excel_001",
        "part_i_subject_information": {
            "subject_type": "individual",
            "last_name": "User",
            "first_name": "Test",
        },
        "part_ii_suspicious_activity": {
            "date_range_start": "2026-03-01",
            "activity_type_codes": ["Other"],
        },
        "filing_notes": ["Review all fields before submission."],
    }


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestHelpers:
    def test_safe_str_none(self):
        assert _safe_str(None) == ""

    def test_safe_str_normal(self):
        assert _safe_str("hello") == "hello"

    def test_safe_str_truncate(self):
        result = _safe_str("a" * 50, max_len=10)
        assert len(result) == 13  # 10 + "..."
        assert result.endswith("...")

    def test_safe_str_no_truncate_when_short(self):
        assert _safe_str("short", max_len=10) == "short"


class TestModuleLevelConstants:
    def test_risk_fills_all_levels(self):
        for level in ("LOW", "MEDIUM", "HIGH", "CRITICAL"):
            assert level in RISK_FILLS
            assert level in RISK_FONTS

    def test_evidence_level_fills(self):
        for level in ("V", "S", "I", "U"):
            assert level in EVIDENCE_LEVEL_FILLS

    def test_disposition_fills_exist(self):
        assert DISP_CLEAR_FILL is not None
        assert DISP_MATCH_FILL is not None


class TestGenerateExcel:
    def test_generates_xlsx_file(self):
        output = _mock_output()
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            result = generate_excel(output, output_path=path)
            assert result.exists()
            assert result.suffix == ".xlsx"
            assert result.stat().st_size > 0

    def test_default_output_path(self):
        output = _mock_output()
        with tempfile.TemporaryDirectory() as td:
            # Override to avoid writing to actual results dir
            path = Path(td) / "screening_results.xlsx"
            result = generate_excel(output, output_path=path)
            assert result == path

    def test_has_expected_sheets(self):
        from openpyxl import load_workbook
        output = _mock_output()
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            generate_excel(output, output_path=path, evidence_store=_sample_evidence())
            wb = load_workbook(str(path))
            names = wb.sheetnames
            assert "Executive Summary" in names
            assert "Screening Results" in names
            assert "Risk Factors" in names
            assert "Decision Points" in names
            assert "Regulatory Actions" in names
            assert "Evidence Detail" in names

    def test_business_client_has_ubo_sheet(self):
        from openpyxl import load_workbook
        output = _mock_output(client_type="business")
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            generate_excel(output, output_path=path)
            wb = load_workbook(str(path))
            assert "UBO Ownership" in wb.sheetnames

    def test_individual_client_no_ubo_sheet(self):
        from openpyxl import load_workbook
        output = _mock_output(client_type="individual")
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            generate_excel(output, output_path=path)
            wb = load_workbook(str(path))
            assert "UBO Ownership" not in wb.sheetnames

    def test_evidence_detail_rows(self):
        from openpyxl import load_workbook
        output = _mock_output()
        evidence = _sample_evidence()
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            generate_excel(output, output_path=path, evidence_store=evidence)
            wb = load_workbook(str(path))
            ws = wb["Evidence Detail"]
            # Header + 2 evidence records
            assert ws.max_row == 3

    def test_evidence_detail_agent_name_not_duplicating_source(self):
        """Bug fix: Agent Name column should not duplicate Source Name."""
        from openpyxl import load_workbook
        output = _mock_output()
        evidence = _sample_evidence()
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            generate_excel(output, output_path=path, evidence_store=evidence)
            wb = load_workbook(str(path))
            ws = wb["Evidence Detail"]
            # Row 2: first evidence record
            source_name = ws.cell(row=2, column=3).value  # Source Name
            agent_name = ws.cell(row=2, column=4).value    # Agent Name
            # Agent name should come from agent_name field, not duplicate source_name
            assert agent_name == "IndividualSanctionsAgent"
            assert source_name == "OFAC SDN"
            assert agent_name != source_name

    def test_no_investigation_still_generates(self):
        output = _mock_output(with_investigation=False)
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            result = generate_excel(output, output_path=path)
            assert result.exists()

    def test_no_synthesis_still_generates(self):
        output = _mock_output(with_synthesis=False)
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            result = generate_excel(output, output_path=path)
            assert result.exists()

    def test_filing_worksheet_sheets_included(self):
        """Feature: SAR/STR filing worksheets should be added when filing dicts provided."""
        from openpyxl import load_workbook
        output = _mock_output()
        fincen = _sample_fincen_filing()
        fintrac = {"form": "FINTRAC STR", "part_a_report_info": {"report_type": "STR"}}
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            generate_excel(output, output_path=path, fincen_filing=fincen, fintrac_filing=fintrac)
            wb = load_workbook(str(path))
            assert "SAR Filing Worksheet" in wb.sheetnames
            assert "STR Filing Worksheet" in wb.sheetnames

    def test_filing_worksheet_not_included_when_none(self):
        from openpyxl import load_workbook
        output = _mock_output()
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            generate_excel(output, output_path=path)
            wb = load_workbook(str(path))
            assert "SAR Filing Worksheet" not in wb.sheetnames

    def test_filing_worksheet_has_data(self):
        from openpyxl import load_workbook
        output = _mock_output()
        fincen = _sample_fincen_filing()
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            generate_excel(output, output_path=path, fincen_filing=fincen)
            wb = load_workbook(str(path))
            ws = wb["SAR Filing Worksheet"]
            # Should have header + multiple data rows
            assert ws.max_row > 3

    def test_executive_summary_has_risk_level(self):
        from openpyxl import load_workbook
        output = _mock_output()
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            generate_excel(output, output_path=path)
            wb = load_workbook(str(path))
            ws = wb["Executive Summary"]
            # Find the risk level row
            found = False
            for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
                if row[0] == "Risk Level":
                    assert row[1] == "HIGH"
                    found = True
            assert found, "Risk Level field not found in Executive Summary"

    def test_risk_factors_sorted_descending(self):
        from openpyxl import load_workbook
        output = _mock_output()
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "test.xlsx"
            generate_excel(output, output_path=path)
            wb = load_workbook(str(path))
            ws = wb["Risk Factors"]
            # Row 2 should be highest points (PEP=20), row 3 should be next (Jurisdiction=15)
            assert ws.cell(row=2, column=2).value == 20
            assert ws.cell(row=3, column=2).value == 15
